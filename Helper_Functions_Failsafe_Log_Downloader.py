import shutil
from bs4 import BeautifulSoup
import webbrowser
from pathlib import Path
from urllib.parse import urlparse
from rich import print as print_
from rich.progress import Progress, BarColumn, TextColumn, TimeRemainingColumn, TimeElapsedColumn
import shutil
import sys
from pynput.keyboard import Controller, Key
import time, re, os
from tkinter import ttk, filedialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --------------------------------------------------------- CONFIG ---------------------------------------------------------
DOWNLOAD_DIR = Path.home() / "Downloads"
PASSED_DIR = Path.home() / "Documents/PASSED"
FAILED_DIR = Path.home() / "Documents/FAILED"
HTML_REPORT_PATH = Path.home() / "Documents/Report.html"
KEYWORD = "measurement"
#KEYWORD = "measurement_xcp"


# ------------------------------------------------------ RICH FORMATTER ------------------------------------------------------
def gen(text: str, style: str = 'bold'):
    """Generate styled rich text"""
    return f"[{style}]{text}[/{style}]"

# ---------------------------------------------------------------------------------------------------------------------------------------

def validate_FAME_url(url: str) -> bool:
    """
    Validates if the given URL matches the expected pattern.
    The URL should contain 'x' in the path and 'report_' in the report segment.
    Example: https://example.com/report/x47/report_2025-05-27T205857%2B0530
    Returns True if valid, False otherwise.
    """
    try:
        # Define the regex pattern
        pattern = r'^https://[a-zA-Z0-9-]+\.conti\.de/hil_status/report/x[0-9]+/report_[0-9]{4}-[0-9]{2}-[0-9]{2}T[0-9]{6}%2B[0-9]{4}$'
        
        # Check if the URL matches the pattern
        if not re.match(pattern, url):
            print_(gen("Invalid URL format! Ensure it contains 'x' followed by digits and 'report_' followed by a valid timestamp.", 'bold #ff471a'))
            return False
        
        # Additional checks for 'x' and 'report_' presence
        if 'x' not in url or 'report_' not in url:
            print_(gen("URL must contain 'x' and 'report_' in the correct segments!", 'bold #ff471a'))
            return False
        
        print_(gen("URL validation successful!", 'bold #00ff00'))
        return True

    except Exception as e:
        print_(gen(f"Error occurred during URL validation: {str(e)}", 'bold #ff471a'))
        return False

# ---------------------------------------------------------------------------------------------------------------------------------------

def extract_links_from_html(html_path: Path, keyword: str) -> list[str]:
    """Extract all anchor tag links from an HTML file that contain a specific keyword."""

    try:
        with html_path.open('r', encoding='utf-8') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            links_data_list = soup.find_all('a') # fetching the anchor tags.

        #print(links_data_list)
        return [link.get('href') for link in links_data_list if keyword in str(link)]
    
    
    except Exception as e:
        print_(gen(f"Error occurred while reading the HTML report!\n{repr(e)}\nPlease try again!", 'bold #ff471a'))
        exit()

# ---------------------------------------------------------------------------------------------------------------------------------------

def segregate_links(links: list[str]):
    """Segregates the links based on the test result"""

    marker = gen(f">>> ", "bold #FDCE01")
    links_dict = {"passed_testcases_links": [], "failed_testcases_links": [], "other_testcases_links": [], "PASSED" : 0, "FAILED" : 0, "OTHERS" : 0}
    for link in links:
        if "PASS" in link:
            links_dict["passed_testcases_links"].append(link)
        elif "FAIL" in link:
            links_dict["failed_testcases_links"].append(link)
        else:
            links_dict["other_testcases_links"].append(link)
    
    links_dict["PASSED"] = len(links_dict["passed_testcases_links"])
    links_dict["FAILED"] = len(links_dict["failed_testcases_links"])
    links_dict["OTHERS"] = len(links_dict["other_testcases_links"])

    print_(f"{gen(f"{marker} PASSED Test Cases: ", "bold " )}{links_dict["PASSED"]}")
    print_(f"{gen(f"{marker} FAILED Test Cases: ", "bold " )}{links_dict["FAILED"]}")
    print_(f"{gen(f"{marker} Other Test Cases: ", "bold " )}{links_dict["OTHERS"]}")
    
    return links_dict

# ---------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------


def z_FAME_downloader(urls: list) -> tuple[int, int, int]:
    """
        Download .mf4 files from a list of URLs with an overall progress bar and error handling.

        Args:
            urls (list): List of URLs, each expected to trigger an .mf4 file download.

        Returns:
            tuple: (downloaded_count, skipped_count, error_count)
    """

    # def gen(text: str, style: str = 'bold') -> str:
    #     """Generate styled rich text."""
    #     return f"[{style}]{text}[/{style}]"
    
    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 

    def extract_filename(url: str) -> str:
        """Extract the filename from a URL."""
        parsed = urlparse(url)
        path = parsed.path
        filename = path.split('/')[-1].split('?')[0]
        return filename

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 

    def is_valid_mf4_url(url: str) -> bool:
        """Check if the URL contains '.mf4'."""
        return '.mf4' in url

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 

    def file_exists(download_dir: Path, filename: str) -> bool:
        """Check if the file exists in the download directory."""
        return (download_dir / filename).is_file()

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 

    def wait_for_file(download_dir: Path, filename: str, timeout: float = 60.0) -> bool:
        """Wait for a file to appear in the download directory."""
        start_time = time.time()
        while time.time() - start_time < timeout:
            if file_exists(download_dir, filename):
                return True
            time.sleep(0.5)
        return False

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 

    marker = gen(f">>> ", "bold #FDCE01")
    download_dir = Path.home() / "Downloads"
    downloaded_count = 0
    skipped_count = 0
    error_count = 0

    # Overall progress bar
    with Progress(
        TextColumn("[magenta]{task.description}"), #TextColumn("[magenta]{task.description}")
        BarColumn(bar_width=None),
        "[progress.percentage]{task.percentage:>3.0f}%",
        TimeElapsedColumn(),
        TimeRemainingColumn(),
        console=None
    ) as overall_progress:
        overall_task = overall_progress.add_task("\n[magenta]Processing all files...", total=len(urls))

        for i, url in enumerate(urls, 1):
            print_(gen(f"\n{marker} Processing URL {i}/{len(urls)}: {url}", "bold cyan"))
            try:
                if not is_valid_mf4_url(url):
                    print_(gen(f"{marker} Skipping URL: {url} - no .mf4 found", "bold #ff471a"))
                    skipped_count += 1
                    continue

                filename = extract_filename(url)
                print_(filename)

                # Check if file already exists
                if file_exists(download_dir, filename):
                    print_(gen(f"\n{marker} File {filename} already exists, skipping.", "bold #49F731"))
                    skipped_count += 1
                    continue

                # Open URL in browser
                webbrowser.open(url)

                # Wait for file to appear
                if wait_for_file(download_dir, filename):
                    print_(gen(f"[+] Successfully downloaded {filename}", "bold #49F731"))
                    downloaded_count += 1
                else:
                    raise TimeoutError(f"File {filename} not detected within 60 seconds 😶")

            except Exception as e:
                print_(gen(f"Error processing {url}: {repr(e)}", "bold #ff471a"))
                error_count += 1
                overall_progress.advance(overall_task)
                overall_progress.refresh()
                time.sleep(0.1)  # Brief delay to ensure progress bar updates
                return downloaded_count, skipped_count, error_count

            overall_progress.advance(overall_task)
            overall_progress.refresh()  # Force refresh to update display

            # Update progress bar color to green when complete
            if overall_progress.tasks[overall_task].completed == len(urls):
                overall_progress.update(overall_task, description="[green]Processing all files...")

    # Print summary
    summary = f"Completed: \n{marker} {downloaded_count} files downloaded \n{marker} {skipped_count} files skipped \n{marker} {error_count} errors"
    #print_(gen(summary, "bold #49F731" if error_count == 0 else "bold #ff471a"))
    print_(gen(summary, "bold #ff471a" if error_count != 0 else "bold"))

    return downloaded_count, skipped_count, error_count

# ---------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------

def move_files(source_path, destination_path: str) -> tuple[int, int]:
    """
    Move .mf4 files from source_path to destination_path with progress bar and error handling.

    Args:
        source_path (str): Path to the source directory containing .mf4 files.
        destination_path (str): Path to the destination directory.

    Returns:
        tuple: (moved_count, error_count)
    """
    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    def validate_path(path: str, is_source: bool = True) -> Path:
        """Validate if a path exists and is a directory."""
        try:
            path_obj = Path(path)
            if not path_obj.exists():
                raise FileNotFoundError(f"{'Source' if is_source else 'Destination'} path does not exist: {path}")
            if not path_obj.is_dir():
                raise NotADirectoryError(f"{'Source' if is_source else 'Destination'} path is not a directory: {path}")
            return path_obj
        except Exception as e:
            print_(gen(f"Error: {repr(e)}", "bold #ff471a"))
            sys.exit(1)
    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    def get_mf4_files(source_path: Path) -> list:
        """Get a list of .mf4 files in the source directory."""
        try:
            return list(source_path.glob("*.mf4"))
        except Exception as e:
            print_(gen(f"Error accessing source directory {source_path}: {repr(e)}", "bold #ff471a"))
            sys.exit(1)

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

    # Validate paths
    src_path = validate_path(source_path, is_source=True)
    dst_path = validate_path(destination_path, is_source=False)

    # Get .mf4 files
    mf4_files = get_mf4_files(src_path)
    if not mf4_files:
        print_(gen(f"No .mf4 files found in {source_path}", "bold #ff471a"))
        return 0, 0

    moved_count = 0
    error_count = 0

    # Overall progress bar
    with Progress(
        TextColumn("[magenta]{task.description}"),
        BarColumn(bar_width=None),
        "[progress.percentage]{task.percentage:>3.0f}%",
        TimeElapsedColumn(),
        TimeRemainingColumn(),
        console=None
    ) as progress:
        task = progress.add_task("[magenta]Moving .mf4 files...", total=len(mf4_files))

        for file in mf4_files:
            try:
                dst_file = dst_path / file.name
                #print_(gen(f"Moving {file.name} to {dst_path}", "bold"))
                shutil.move(str(file), str(dst_file))
                print_(gen(f"[+] Successfully moved {file.name}", "bold #49F731"))
                moved_count += 1
            except Exception as e:
                print_(gen(f"Error moving {file.name}: {repr(e)}", "bold #ff471a"))
                error_count += 1
                progress.advance(task)
                return moved_count, error_count  # Exit on any error
            progress.advance(task)
            progress.refresh()

        # Update progress bar to green when complete
        if progress.tasks[task].completed == len(mf4_files):
            progress.update(task, description="[green]Moving .mf4 files...")

    # Print summary
    summary = f"Completed: \n{moved_count} files moved, \n{error_count} errors"
    print_(gen(summary, "bold #ff471a" if error_count != 0 else "bold"))

    return moved_count, error_count

# ---------------------------------------------------------------------------------------------------------------------------------------

def HTML_Downloader(url: str):
    """
    Opens a URL in the default browser, simulates Ctrl+S and Enter to save the HTML
    after JavaScript rendering, and returns the path of the saved file.

    Args:
        url (str): The URL of the webpage to open and save.
        download_dir (str): Directory where the HTML file is saved (default: ~/Downloads).

    Returns:
        str: Path to the saved HTML file, or empty string on failure.
    """

    def validate_url(url: str) -> bool:
        """Validate if the URL is well-formed."""
        try:
            if not url.startswith(('http://', 'https://')):
                raise ValueError("URL must start with http:// or https://")
            return True
        except Exception as e:
            print_(gen(f"Invalid URL {url}: {repr(e)}", "bold #ff471a"))
            return False
    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 
    
    def find_newest_html_file(download_dir: Path, max_attempts: int = 6, delay: float = 2.0) -> str:
        """Find the newest .html file in the download directory with retries."""
        try:
            for _ in range(max_attempts):
                html_files = list(download_dir.glob("*.html"))
                if html_files:
                    newest_file = max(html_files, key=lambda p: p.stat().st_mtime)
                    return str(newest_file)
                time.sleep(delay)
            return ""
        except Exception as e:
            print_(gen(f"Error finding HTML file in {download_dir}: {repr(e)}", "bold #ff471a"))
            return ""
    
    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 

    def extract_filename(url: str) -> str:
            """Extract the filename from a URL."""
            parsed = urlparse(url)
            path = parsed.path
            filename = path.split('/')[-1].split('?')[0]
            return filename

    # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 

    download_dir: Path = Path.home() / "Downloads"
    report_file_name = extract_filename(url)

    # Validate URL
    if not validate_url(url):
        return ""
    
    # Warn about keyboard interference
    print_(gen("Warning: Do not use keyboard/mouse while script saves HTML", "bold #ff471a"))

    # Progress bar for loading and saving ------------------------------------------------------------------------
    with Progress(
        TextColumn("[magenta]{task.description}"),
        BarColumn(bar_width=None),
        console=None,
        transient=True
        ) as progress:
        
        task = progress.add_task(f"[magenta]Opening & saving {report_file_name}...", total=1)

        try:
            # Open URL in browser
            webbrowser.open(url)
            
            # Wait for page to load and JavaScript to render
            time.sleep(5)  # Increased delay for JS rendering
            # if not check_page_load(url):
            #     return ""

            # Keyboard automation
            keyboard = Controller()
            # Ctrl+S to open Save As dialog
            with keyboard.pressed(Key.ctrl):
                keyboard.type('s')  # Using your suggested type() for simplicity
            time.sleep(1)  # Wait for dialog to open

            # Alt+D to focus address bar
            with keyboard.pressed(Key.alt):
                keyboard.type('d')
            time.sleep(0.5)  # Wait for address bar focus

            # Type Downloads path and Enter
            keyboard.type(str(download_dir))
            time.sleep(0.5)  # Brief pause

            # Press enter key 5 times
            for _ in range(4):
                keyboard.type('\n')
                time.sleep(0.5)  # Brief pause

            # Wait and find the newest HTML file
            saved_file = find_newest_html_file(download_dir)
            if not saved_file:
                print_(gen(f"No HTML file found in {download_dir} after saving", "bold #ff471a"))
                return ""

            print_(gen(f"[+] Successfully downloaded {url} to {saved_file}", "bold green"))
            progress.advance(task)
            return Path(saved_file)

        except Exception as e:
            print_(gen(f"Error processing {url}: {repr(e)}", "bold #ff471a"))
            return ""
# ---------------------------------------------------------------------------------------------------------------------------------------

def get_HTML_path():
    try:
        print_("\n{}{}".format(gen('---> ', 'bold #FEFA02'), gen("Press Enter key if you want to automatically download the HTML file.", 'bold #00ffff')))
        print_("{}{}\n".format(gen('---> ', 'bold #FEFA02'), gen("Press '+' key & enter key to manually select the HTML file.", 'bold #00ffff')))
        ch = input('Enter your choice or path : ')
        print('\n')

        # Case 1 - when path.txt is present & the user enters just a blank space or hits enter
        if (ch in ["", ' ',]):
                FAME_REPORT_URL : str = input("\nEnter the URL - ")
                return HTML_Downloader(FAME_REPORT_URL)
        # Case 2
        elif (ch == "+"):
                def path_resolver(path:str):
                    from pathlib import Path
                    return Path(path.replace('\\', "/"))

                path = filedialog.askopenfilename()
                path = path_resolver(path)
                if str(path).strip().endswith('html'):
                      return path
                else:
                      print_(gen("Invalid file!\nPlease select a HTML file.", 'bold #ff471a')) 
                      
        else:
             print_(gen("Invalid option!", 'bold #ff471a'))
             exit()

    except Exception as e:
        print_(gen(f"Error {repr(e)} occured fetching HTML file.\nPlease try again!", 'bold #ff471a'))
        exit()

# ---------------------------------------------------------------------------------------------------------------------------------------   
#---------------------------------------------------------------------------------------------------------------------------------------
def TR_path_fetch():
    try:
        current_dir = os.getcwd() # fetches the path of the current working directory
        current_dir_contents = os.listdir(current_dir) # stores all the files & folders present in the current working directory
        print_("{}{}\n".format(gen('---> ', 'bold #FEFA02'), gen("Press Enter Key to manually select the TR.", 'bold #00ffff')))
        path = filedialog.askopenfilename()
        if path[-4:] == "xlsx":
            print_(f"[bold #00ff00]{"\nTR path fetch successful !"}[/bold #00ff00]")
            #print_(f"{gen("\nReading Data from DMD...", "#A7FF03")}")
            return path
        else:
             print_(gen("Invalid path!", 'bold #ff471a'))
             exit()

    except:
        print_(gen("Error occured while reading TR!\nPlease try again!", 'bold #ff471a'))
        exit()
#---------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------   

def TR_update(TR_path : str, logs_path):

    # === USER CONFIGURATION ===

    excel_path = TR_path
    sheet_name = "Test Cases - Test Log"

    # === STEP 1: Load Workbook and Sheet ===
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # === STEP 2: Get Column Letters Based on Headers ===
    header_row = 3
    col_mapping = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col).value
        if header_val:
            col_mapping[str(header_val).strip()] = get_column_letter(col).strip()
        #print(header_val, "\n", "---"*36)

    col_test_id = col_mapping.get("Automated \nTest Case ID")
    col_measure = col_mapping.get("Measurements / Time\n(Filename, picture)")
    col_result = col_mapping.get("Result")
    col_deviation = col_mapping.get("Deviation Description \n(Comment from Tester)")

    col_auto_eval_result = col_mapping.get("Automated Evaluation Result")
    col_on_hold = col_mapping.get("On Hold")
    col_on_hold_comment = col_mapping.get("On Hold Comment")


    
    if all([col_test_id, col_measure, col_result, col_deviation, col_auto_eval_result, col_on_hold, col_on_hold_comment]):
        print(f"[+] Column data updated correctly 👍")
    else:
        raise ValueError("One or more required columns were not found.")

    # === STEP 3: Get List of Log Files ===
    log_files = os.listdir(logs_path)

    # === STEP 4: Iterate Rows and Update ===
    row = header_row + 1
    empty_counter = 0
    max_empty_allowed = 18

    while empty_counter < max_empty_allowed:
        test_id_cell = ws[f"{col_test_id}{row}"]
        test_id = test_id_cell.value.strip() if test_id_cell.value else ""
        
        if not test_id:
            empty_counter += 1
            row += 1
            continue
        empty_counter = 0
        test_id_lower = test_id.lower()

        match_found = False
        for log_file in log_files:
            log_lower = log_file.lower()
            # checking if the 
            if test_id_lower in log_lower:
                ws[f"{col_measure}{row}"].value = log_file
                if "pass" in log_lower:
                    ws[f"{col_result}{row}"].value = "PASSED"
                    ws[f"{col_auto_eval_result}{row}"].value = "FINAL"
                    ws[f"{col_on_hold}{row}"].value = "no"
                
                else:
                    ws[f"{col_result}{row}"].value = "FAILED"
                    ws[f"{col_deviation}{row}"].value = "FAILED"
                match_found = True
                break

        if not match_found:
            #ws[f"{col_measure}{row}"].value = "EMPTY"
            pass

        row += 1

    # === STEP 5: Save Changes ===
    wb.save(f"{TR_path[:-5]}_Updated.xlsx")
    print("Update complete.")


# ---------------------------------------------------------------------------------------------------------------------------------------   
